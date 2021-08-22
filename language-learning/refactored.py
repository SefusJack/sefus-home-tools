import urllib.request, json
from urllib.parse import quote 
import requests 
import openpyxl
from openpyxl import Workbook
from bs4 import BeautifulSoup
import re
import time

wordstash = list()
kanjistash = list()
kanjidata = list()

pound="%23"
n5=pound+"jlpt-n5"
n4=pound+"jlpt-n4"
n3=pound+"jlpt-n3"
n2=pound+"jlpt-n2"
n1=pound+"jlpt-n1"
definition=""

english = ["Ａ", "Ｂ", "Ｃ", "Ｄ", "Ｅ", "Ｆ", "Ｇ", "Ｈ", "Ｉ", 
            "Ｊ", "Ｋ", "Ｌ", "Ｍ", "Ｎ", "Ｏ", "Ｐ", "Ｑ", "Ｒ", "Ｓ", "Ｔ", "Ｕ", "Ｖ", "Ｗ",
            "Ｘ", "Ｙ", "Ｚ", "０", "１", "２", "３", "４", "５", "６", "７", "８", "９"]

try:
    wb = openpyxl.load_workbook("sample.xlsx")
    ws = wb["Sheet"]
    ws2 = wb["Kanji"]
except:
    wb = Workbook()
    ws = wb.active
    ws2 = wb.create_sheet("Kanji")
    ws.append(["Word", "Word Audio", "Original Word", "Reading", "Meaning", "Sentence", "Sentence Audio", "Sentence Meaning", "Parts of Speech" , "Kanji 1", "Meaning 1", "Kunyomi 1", "Onyomi 1", "Kanji 2", "Meaning 2", "Kunyomi 2", "Onyomi 2", "Kanji 3" ,"Meaning 3", "Kunyomi 3", "Onyomi 3", "Kanji 4", "Meaning 4", "Kunyomi 4", "Onyomi 4", "Picture", "Tags"])
    ws2.append(["Kanji", "Meaning", "Frequency", "Kunyomi", "Onyomi"])

#Returns a 2D list of all the Rows data from an excel worksheet
def getDataFromAllExcelRows(worksheet, header):
    rowsdata = list()
    for i in range(worksheet.min_row+header, worksheet.max_row):
        rowsdata.append(getDataFromExcelRow(worksheet, i))
    return rowsdata

#Returns a list of the data from an inputted row and excel worksheet
def getDataFromExcelRow(worksheet, row):
    rowdata = list()
    for col in worksheet[row]:
        rowdata.append(col.value)
    return rowdata

#Returns a 2D list of all the Columns data from an excel worksheet
def getDataFromAllExcelColumns(worksheet, header):
    colsdata = list()
    for i in range(worksheet.min_column, worksheet.max_column):
        colsdata.append(getDataFromExcelColumn(worksheet, header, i-1))
    return colsdata

#Returns a list of the data from an inputted column and excel worksheet
def getDataFromExcelColumn(worksheet, header, column):
    coldata = list()
    for row in worksheet.iter_rows(worksheet.min_row+header,worksheet.max_row):
        coldata.append(row[column].value)
    return coldata

#Returns the data from a single cell from a worksheet
def getDataFromCell(worksheet, row, column):
    return worksheet.cell(row, column).value

#Returns True if the inputted character is a kanji, Returns false otherwise
def isKanji(character):
    #々 is an iteration mark
    #ヶ can be considered an abbreviation
    return re.match(r'([^一-龯])', character) is None or character == "々" or character == "ヶ"

#Returns a list of kanji found in a word
def getKanjiFromWord(word):
    temp = list()
    if word is not None:
        for character in word:
            if isKanji(character):
                temp.append(character)
    return temp

#Returns a list of kanji from a list of words
def getKanjiListFromWords(words):
    kanjistemp = list()
    for word in words:
        kanjis = getKanjiFromWord(word)
        for kanji in kanjis:
            if kanji not in kanjistemp:
                kanjistemp.append(kanji)
    return kanjistemp

#Returns a list of kanji from a list of words (including only kanji) that are currently in the inputted excel worksheet, header number, and column to search in
def getKanjiListFromExcelWorkSheet(worksheet, header, column):
    return getKanjiListFromWords(getDataFromExcelColumn(worksheet, header, column))

#Returns true if the Kanji inputted is found in the Kanji stash, returns false otherwise
def inKanjiStash(kanji):
    return kanji in kanjistash

#Adds a List of Kanji to the Kanji Stash
def addKanjiListToStash(kanjis):
    for kanji in kanjis:
        if not inKanjiStash(kanji):
            kanjistash.append(kanji)

#Adds a Kanji to the Kanji stash
def addKanjiToStash(kanji):
    if not inKanjiStash(kanji):
        kanjistash.append(kanji)

#Adds all of the Kanji from the Kanji Stash To Excel. Kanji already in the Excel WorkSheet will be ignored
def addKanjiStashToExcel(worksheet, header, column):
    for kanji in kanjistash:
        temp = getKanjiListFromExcelWorkSheet(worksheet, header, column)
        if kanji not in temp:
            worksheet.append([kanji])

def getKanjiDataFromKanjiStash():
    for kanji in kanjistash:
        kanjidata.append(kanjiSearch(kanji))

###########################################################################################################
def getWord(data):
    if 'word' in data['japanese'][0]:
        return data['japanese'][0]['word']
    else:
        return getReading(data)

def getReading(data):
    if 'reading' in data['japanese'][0]:
        return data['japanese'][0]['reading']
    return ""

def getMeaning(data, limit, multiwordf):
    temp = ""
    wordcharpos = 1
    length = len(data['senses'])
    if data['senses'][length-1]['parts_of_speech']:
        if data['senses'][length-1]['parts_of_speech'][0] == 'Wikipedia definition':
            length -= 1
    for i in data['senses']:
            if wordcharpos < length and wordcharpos < limit:
                temp = temp + '"' + str(wordcharpos) + ". " + ", ".join(i['english_definitions']).replace('"','') + '"' + " &CHAR(10)& "
            else:
                if not multiwordf:
                    temp = temp + '"' + str(wordcharpos) + ". " + ", ".join(i['english_definitions']).replace('"','') + '"'
                    return "= " + temp
                else:
                    return ", ".join(i['english_definitions']).replace('"','')
            wordcharpos += 1

def getJLPT(data):
    if data['jlpt']:
        return list(reversed(sorted(data['jlpt'])))[0]
    else:
        return ""

def getCommonality(data):
    if data['is_common'] != None:
        if data['is_common'] == True:
            return "common"
        else:
            return "uncommon"
    else:
        return ""

def getPartOfSpeech(data):
    temp = []
    for i in data['senses'][0]['parts_of_speech']:
        if i == "Suru verb" or i == "Godan verb with 'ru' ending" or i == "Intransitive verb" or i == "Ichidan verb" or i == "Transitive verb" or i == "Godan verb with 'mu' ending" or i == "Godan verb with 'su' ending" or i == "Godan verb with 'u' ending" or i == "Godan verb with 'ku' ending" or i == "Kuru verb - special class" or i == "Suru verb - included" or i == "Godan verb with 'bu' ending" or i == "Godan verb with 'gu' ending" or i == "Noun or verb acting prenominally" or i == "Godan verb with 'nu' ending" or i == "Irregular nu verb" or i == "Godan verb - Iku/Yuku special class" or i == "Godan verb with 'tsu' ending" or i == "Godan verb with 'ru' ending (irregular verb)" or i == "Ichidan verb - kureru special class" or i == "Godan verb - -aru special class" or i == "Auxiliary verb":
            temp.append("Verb")
        elif i == "Adverb (fukushi)" or i == "Adverb taking the 'to' particle":
            temp.append("Adverb")
        elif i == "Na-adjective (keiyodoshi)" or i == "I-adjective (keiyoushi)" or i == "Pre-noun adjectival (rentaishi)":
            temp.append("Adjective")
        elif i == "Noun which may take the genitive case particle 'no'" or i == "Noun, used as a prefix" or i == "Noun, used as a suffix":
            temp.append("Noun")
        elif i == "Expressions (phrases, clauses, etc.)":
            temp.append("Expressions")
        else:
            temp.append(i)
        
    temp = list(dict.fromkeys(temp))
    return  " ".join(temp)

def getKanji(word):
    temp = list()
    for i in list(word):
        if i not in hiragana and i not in katakana and i not in halfwidth and i not in english:
            temp.append(i)
            if i not in kanji:
                kanji.append(i)
                kanjidata.append(kanjiSearch(i))
    return temp

def isEnglish(word):
    for i in list(word):
        if i in english:
            return True
    return False

def isOnlyKanji(word):
    for i in list(word):
        if not isKanji(i):
            return False
    return True

def hasKanji(word):
    for i in list(word):
        if isKanji(i):
            return True
    return False

def getFurigana(word, reading):
    if not isEnglish(word) and word != "":
        furiword = ""
        if isOnlyKanji(word):
            return word + "[" + reading + "]"
        else:
            prevnode = None
            currentnode = None
            nextnode = None
            wordpos = 0
            readpos = 0
            wordpos != len(word)-1
            while wordpos < len(word) and readpos < len(reading):
                currentnode = word[wordpos]
                if wordpos != len(word)-1:
                    nextnode = word[wordpos+1]
                else:
                    nextnode = None
                
                if isKanji(currentnode):
                    furiword = furiword + currentnode
                    if nextnode == None:
                        furiword = furiword + "[" + reading[readpos:] + "]"
                    wordpos += 1
                    prevnode = currentnode
                else:
                    if prevnode == None:
                        furiword = furiword + currentnode
                        prevnode = currentnode
                        readpos += 1
                        wordpos += 1
                    else:
                        if isKanji(prevnode):
                            if nextnode == None:
                                furiword = furiword + "["
                                while readpos != len(reading)-1:
                                    furiword = furiword + reading[readpos]
                                    readpos += 1
                                furiword = furiword + "]" + currentnode
                                readpos = len(reading)
                            else:
                                readpos += 1
                                if isKanji(nextnode):
                                    furiword = furiword + "[" + reading[:readpos]
                                    furiword = furiword + "]"
                                else:
                                    if hasKanji(word[wordpos:]):
                                        furiword = furiword + "["
                                        while not isKanji(word[wordpos]):
                                            furiword = furiword + word[wordpos]
                                            wordpos += 1
                                            readpos += 1
                                        furiword = furiword + "]"
                                        readpos += 1
                                    else:
                                        furiword = furiword + "[" + reading[:readpos]
                                        while reading[readpos:] != word[wordpos:]:
                                            furiword = furiword + reading[readpos]
                                            readpos += 1
                                        furiword = furiword + "]" + reading[readpos:]
                                        readpos += len(reading)
                                prevnode = currentnode
                        else:
                            furiword = furiword + currentnode
                            prevnode = currentnode
                            wordpos += 1
                            readpos += 1
        return furiword
    else:
        return reading

def request(page, search):
    with urllib.request.urlopen("https://jisho.org/api/v1/search/words?keyword=" + str(search) + "&page=" + str(page)) as site:
            time.sleep(0.5)
            result = json.loads(site.read().decode())
    return result

def readPage(start, limit, search):
        page = start
        if limit == 0:
            limit = sys.maxint
        data = list()
        while(page <= limit):
            result = request(page, search)
            if len(result['data']) == 0:
                return data
            else:
                data.append(result)
            page += 1
        return data

def wordSearchToExcel(worksheet, header, search):
    print("Searching For:" + removeFurigana(search))
    data = readPage(1, 1, quote(removeFurigana(search)))
    if data:
        j = data[0]['data'][0]
        temp = list()
        word = getWord(j)
        reading = getReading(j)
        meaning = getMeaning(j, 3, False)
        kanjiused =  getKanjiFromWord(word)
        furigana = getFurigana(word, reading)
        partsofspeech = getPartOfSpeech(j)
        tags = [getJLPT(j),getCommonality(j), partsofspeech]
        try:
            kanji1 = kanjiused[0]
        except:
            kanji1 = ""
        
        try:
            kanji2 = kanjiused[1]
        except:
            kanji2 = ""
        
        try:
            kanji3 = kanjiused[2]
        except:
            kanji3 = ""
        
        try:
            kanji4 = kanjiused[3]
        except:
            kanji4 = ""
        
        temp = [furigana, "", word, reading, meaning, "", "", "",  partsofspeech.replace(' ',', '), kanji1, "", "", "", kanji2, "", "", "", kanji3, "", "", "", kanji4, "", "", "", "", " ".join(tags)]
        if search in wordstash:
            for i in range(worksheet.min_column, worksheet.max_column+1):
                worksheet.cell(wordstash.index(search)+header+1, i).value = temp[i-1]
            wordstash[wordstash.index(search)] = furigana
        elif furigana in wordstash:
            for i in range(worksheet.min_column, worksheet.max_column+1):
                worksheet.cell(wordstash.index(furigana)+header+1, i).value = temp[i-1]
        else:
            wordstash.append(furigana)
            worksheet.append(temp)


def tagSearchToExcel(start, limit, search):
    data = readPage(start, limit, search)
    for i in data:
        for j in i['data']:
            word = getWord(j)
            reading = getReading(j)
            meaning = getMeaning(j, 3, False)
            kanjiused =  getKanjiFromWord(word)
            furigana = getFurigana(word, reading)
            partsofspeech = getPartOfSpeech(j)
            tags = [getJLPT(j),getCommonality(j), partsofspeech]
            if furigana not in wordstash:
                wordstash.append(furigana)
                ws.append([furigana, "", word, reading, meaning, "", "", "",  partsofspeech.replace(' ',', '), "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", " ".join(tags)])

def getListOfKanjiInKanjiData():
    kanjis = list()
    for i in kanjidata:
        kanjis.append(i[0])
    return kanjis

def getwordcharposOfKanjiInKanjiData(kanji):
    return getListOfKanjiInKanjiData().index(kanji)


def kanjiSearch(kanji):
    if kanji not in getListOfKanjiInKanjiData():
        url = requests.get("https://jisho.org/search/" + str(kanji) + "%20%23kanji")
        soup = BeautifulSoup(url.content, 'html.parser')

        frequency = ""
        meaning = ""
        kunyomi = ""
        onyomi = ""
        grade = ""
        jlpt = ""
        #Newspaper Frequency
        if soup.find_all("div", {"class": "frequency"}):
            frequency = soup.find_all("div", {"class": "frequency"})[0].text.strip()
        #Kanji Meaning
        if soup.find_all("div", {"class": "kanji-details__main-meanings"}):
            meaning = soup.find_all("div", {"class": "kanji-details__main-meanings"})[0].text.strip()
        #Kunyomi
        try:
            if soup.find_all("dd", {"class": "kanji-details__main-readings-list"}):
                kunyomi = soup.find_all("dd", {"class": "kanji-details__main-readings-list"})[0].text.strip()
        except:
            pass
        #Onyomi
        try:
            if soup.find_all("dd", {"class": "kanji-details__main-readings-list"}):
                onyomi = (soup.find_all("dd", {"class": "kanji-details__main-readings-list"})[1].text.strip())
        except:
            pass
        
        #Grade
        try:
            if soup.find_all("div", {"class": "grade"}):
                grade = soup.find_all("div", {"class": "grade"})[0].text.strip()
                grade = grade[grade.find('grade'):].replace(' ','-')
        except:
            pass

        #JLPT
        try:
            if soup.find_all("div", {"class": "jlpt"}):
                jlpt = soup.find_all("div", {"class": "jlpt"})[0].text.strip()
                jlpt = "jlpt-" + jlpt[11:].lower()
        except:
            pass
        tags = jlpt + " " + grade
        words = readPage(1, 1, "*" + urllib.parse.quote(str(kanji) + "*"))
        definitions = ""
        wordcharpos = 0
        for j in words:
                for k in j['data']:
                    if wordcharpos < 5:
                        defword = getWord(k)
                        defreading = getReading(k)
                        defmeaning = getMeaning(k, 1, True)
                        deffurigana = getFurigana(defword, defreading)
                        definitions = definitions + '"' + deffurigana + " " + defmeaning + '"' + " &CHAR(10)& "
                        wordcharpos += 1
        definitions = "= " + definitions[:-12]
        info = [kanji, meaning, frequency, kunyomi, onyomi, definitions, tags]
        kanjidata.append(info)
    else:
        info = kanjidata[getwordcharposOfKanjiInKanjiData(kanji)]
    return info
    

def addKanjiDataToWords(worksheet, header, column):
    addKanjiListToStash(getKanjiListFromExcelWorkSheet(worksheet, header, column))
    words = getDataFromExcelColumn(worksheet, header, column)
    currentrow = header+1
    for kanjis in words:
        temp = getKanjiFromWord(kanjis)
        currentcol = 10
        for kanji in temp:
            try:
                data = kanjidata[getwordcharposOfKanjiInKanjiData(kanji)]
            except:
                kanjiSearch(kanji)
                data = kanjidata[getwordcharposOfKanjiInKanjiData(kanji)]

            worksheet.cell(currentrow, currentcol).value = kanji
            currentcol += 1
            worksheet.cell(currentrow, currentcol).value = data[1]
            currentcol += 1
            worksheet.cell(currentrow, currentcol).value = data[3]
            currentcol += 1
            worksheet.cell(currentrow, currentcol).value = data[4]
            currentcol += 1
        currentrow += 1


def kanjiSearchToExcel(kanji):
    for i in kanji:
        if i not in getListOfKanjiInKanjiData():
            info = kanjiSearch(i)
            ws2.append(info)

def getOnyomi(kanji):
        return kanjiSearch(kanji)[4]

def getKunyomi(kanji):
        return kanjiSearch(kanji)[3]

def removeFurigana(word):
    wordcharacters = list(word)
    temp = list()
    ignore = False
    for i in range(0, len(wordcharacters)):
        if wordcharacters[i] == "[":
            ignore = True
        elif wordcharacters[i] == "]":
            ignore = False
        elif ignore == False:
            temp.append(wordcharacters[i])

    return "".join(temp)


#Startup
kanjistash = getKanjiListFromExcelWorkSheet(ws, 1, 0)
print(kanjistash)
kanjidata = getDataFromAllExcelRows(ws2, 1)
wordstash = getDataFromExcelColumn(ws, 1, 0)
#Start Here
#tagSearchToExcel(1, 1000, n1)
#addKanjiListToStash(getKanjiListFromExcelWorkSheet(ws, 1, 0))
#print(wordstash[3])
#getFurigana("ご主人", "ごしゅじん")
#getFurigana("何時も", "いつも")
#getFurigana("喋る", "しゃべる")
#getFurigana("ご飯", "ごはん")
#getFurigana("小さい", "ちいさい")
#getFurigana("食べ物", "たべもの")
#getFurigana("詰まり", "つまり")
#getFurigana("私たち", "わたしたち")
#getFurigana("お願い", "おねがい")
#for i in wordstash:
#    wordSearchToExcel(ws, 1, i)
#print(removeFurigana(wordstash[3]))
#wordSearchToExcel(ws, 1, removeFurigana(wordstash[3]))
kanjiSearchToExcel(kanjistash)
#addKanjiDataToWords(ws, 1, 0)
#End
wb.save("sample.xlsx")