import urllib.request, json
import requests 
import openpyxl
from bs4 import BeautifulSoup

try:
    wb = openpyxl.load_workbook("sample.xlsx")
except:
    wb = Workbook()

#grab the active worksheet
ws = wb.active
ws2 = wb.create_sheet("Kanji")

pound="%23"
n5=pound+"jlpt-n5"
n4=pound+"jlpt-n4"
n3=pound+"jlpt-n3"
n2=pound+"jlpt-n2"
n1=pound+"jlpt-n1"
definition=""

try:
    wordslist = list()
    for row in ws.iter_rows(ws.min_row+1,ws.max_row):
        wordslist.append(row[0].value)
except:
    wordslist = list()

try:
    kanjilist = list()
    for row in ws2.iter_rows(ws2.min_row+1,ws2.max_row):
        kanjilist.append(row[0].value)
except:
    kanjilist = list()

kanjidata = list()

hiragana = ["あ", "い", "う", "え", "お", 
            "か", "き", "く", "け", "こ", 
            "さ", "し", "す", "せ", "そ", 
            "た", "ち", "つ", "て", "と", 
            "な", "に", "ぬ", "ね", "の", 
            "は", "ひ", "ふ", "へ", "ほ", 
            "ま", "み", "む", "め", "も",
            "や",       "ゆ",       "よ",
            "ら", "り", "る", "れ", "ろ",
            "わ",                   "を",
            "が", "ぎ", "ぐ", "げ", "ご",
            "ざ", "じ", "ず", "ぜ", "ぞ",
            "だ", "ぢ", "づ", "で", "ど",
            "ば", "び", "ぶ", "べ", "ぼ"
            "ぱ", "ぴ", "ぷ", "ぺ", "ぽ",
            "ん"]
katakana = ["ア", "イ", "ウ", "エ", "オ",
            "カ", "キ", "ク", "ケ", "コ",
            "ガ", "ギ", "グ", "ゲ", "ゴ",
            "サ", "シ", "ス", "セ", "ソ",
            "ザ", "ジ", "ズ", "ゼ", "ゾ",
            "タ", "チ", "ツ", "テ", "ト",
            "ダ", "ヂ", "ヅ", "デ", "ド",
            "ナ", "ニ", "ヌ", "ネ", "ノ",
            "ハ", "ヒ", "フ", "ヘ", "ホ",
            "バ", "ビ", "ブ", "ベ", "ボ",
            "パ", "ピ", "プ", "ペ", "ポ",
            "マ", "ミ", "ム", "メ", "モ",
            "ヤ",       "ユ",       "ヨ",
            "ラ", "リ", "ル", "レ", "ロ",
            "ワ",                   "ヲ",
            "ン"]
halfwidth = ["ー", "ィ", "ォ", "ッ", "ャ", "ュ", "っ", "ゃ", "々"]
english = ["Ａ", "Ｂ", "Ｃ", "Ｄ", "Ｅ", "Ｆ", "Ｇ", "Ｈ", "Ｉ", 
            "Ｊ", "Ｋ", "Ｌ", "Ｍ", "Ｎ", "Ｏ", "Ｐ", "Ｑ", "Ｒ", "Ｓ", "Ｔ", "Ｕ", "Ｖ", "Ｗ",
            "Ｘ", "Ｙ", "Ｚ", "０", "１", "２", "３", "４", "５", "６", "７", "８", "９"]

#Data can be assigned directly to cells
ws.append(["Word", "Meaning", "Tags", "Parts of Speech", "Kanji 1", "Meaning 1", "Kunyomi 1", "Onyomi 1", "Kanji 2", "Meaning 2", "Kunyomi 2", "Onyomi 2", "Kanji 3", "Meaning 3", "Kunyomi 3", "Onyomi 3", "Kanji 4", "Meaning 4", "Kunyomi 4", "Onyomi 4" ])
ws2.append(["Kanji", "Meaning", "Frequency", "Onyomi", "Kunyomi"])
def getWord(data):
    if 'word' in data['japanese'][0]:
        return data['japanese'][0]['word']
    else:
        return getReading(data)

def getReading(data):
    return data['japanese'][0]['reading']

def getMeaning(data, limit, multiwordf):
    temp = ""
    index = 1
    length = len(data['senses'])
    if data['senses'][length-1]['parts_of_speech']:
        if data['senses'][length-1]['parts_of_speech'][0] == 'Wikipedia definition':
            length -= 1
    for i in data['senses']:
            if index < length and index < limit:
                temp = temp + '"' + str(index) + ". " + ", ".join(i['english_definitions']).replace('"','') + '"' + " &CHAR(10)& "
            else:
                if not multiwordf:
                    temp = temp + '"' + str(index) + ". " + ", ".join(i['english_definitions']).replace('"','') + '"'
                    return "= " + temp
                else:
                    return ", ".join(i['english_definitions']).replace('"','')
            index += 1

def getJLPT(data):
    if data['jlpt']:
        return list(reversed(sorted(data['jlpt'])))[0]
    else:
        return ""

def getCommonality(data):
    if data['is_common'] != None:
        if data['is_common'] == True:
            return "common"
        else:
            return "uncommon"
    else:
        return ""

def getPartOfSpeech(data):
    temp = []
    for i in data['senses'][0]['parts_of_speech']:
        if i == "Suru verb" or i == "Godan verb with 'ru' ending" or i == "Intransitive verb" or i == "Ichidan verb" or i == "Transitive verb" or i == "Godan verb with 'mu' ending" or i == "Godan verb with 'su' ending" or i == "Godan verb with 'u' ending" or i == "Godan verb with 'ku' ending" or i == "Kuru verb - special class" or i == "Suru verb - included" or i == "Godan verb with 'bu' ending" or i == "Godan verb with 'gu' ending" or i == "Noun or verb acting prenominally" or i == "Godan verb with 'nu' ending" or i == "Irregular nu verb" or i == "Godan verb - Iku/Yuku special class" or i == "Godan verb with 'tsu' ending" or i == "Godan verb with 'ru' ending (irregular verb)":
            temp.append("Verb")
        elif i == "Adverb (fukushi)" or i == "Adverb taking the 'to' particle":
            temp.append("Adverb")
        elif i == "Na-adjective (keiyodoshi)" or i == "I-adjective (keiyoushi)" or i == "Pre-noun adjectival (rentaishi)":
            temp.append("Adjective")
        elif i == "Noun which may take the genitive case particle 'no'" or i == "Noun, used as a prefix" or i == "Noun, used as a suffix":
            temp.append("Noun")
        elif i == "Expressions (phrases, clauses, etc.)":
            temp.append("Expressions")
        else:
            temp.append(i)
        
    temp = list(dict.fromkeys(temp))
    return  " ".join(temp)

def getKanji(word):
    temp = list()
    for i in list(word):
        if i not in hiragana and i not in katakana and i not in halfwidth and i not in english:
            temp.append(i)
            if i not in kanjilist:
                kanjilist.append(i)
                kanjidata.append(kanjiSearch(i))
                
    return temp

def isKanji(char):
    if char not in hiragana and not char in katakana and char not in halfwidth and char not in english:
        return True
    return False

def isEnglish(word):
    for i in list(word):
        if i in english:
            return True
    return False

def getFurigana(word, reading):
    wordcharacters = list(word)
    readingcharacters = list(reading)
    furichars = list()
    furiword = ""
    lastkanjipos = -1
    for i in readingcharacters:            
        if i not in wordcharacters:
            furichars.append(i)
    index = 0
    for j in wordcharacters:
        index += 1
        if isKanji(j):
            lastkanjipos = index
        if isEnglish(word):
            return reading
    
    if lastkanjipos > -1:
        return word[:lastkanjipos] + "[" + "".join(furichars) + "]" + word[lastkanjipos:]
    return word

def readPage(start, limit, search):
    page = start
    if limit == 0:
        limit = sys.maxint
    data = list()
    while(page <= limit):
        print("Page Number:" + str(page))
        with urllib.request.urlopen("https://jisho.org/api/v1/search/words?keyword=" + search + "&page=" + str(page)) as site:
            result = json.loads(site.read().decode())
        if len(result['data']) == 0:
            return data
        else:
            data.append(result)
        page += 1
    return data

def wordSearchToExcel(start, limit, search):
    data = readPage(start, limit, search)
    for i in data:
        for j in i['data']:
            word = getWord(j)
            reading = getReading(j)
            meaning = getMeaning(j, 3, False)
            kanjiused = getKanji(word)
            try:
                kanji1 = kanjiused[0]
                meaning1 = kanjidata[kanji.index(kanji1)][0]
                kunyomi1 = kanjidata[kanji.index(kanji1)][1]
                onyomi1 = kanjidata[kanji.index(kanji1)][2]
            except:
                kanji1 = ""
                meaning1 = ""
                kunyomi1 = ""
                onyomi1 = ""
            try:
                kanji2 = kanjiused[1]
                meaning2 = kanjidata[kanji.index(kanji2)][0]
                kunyomi2 = kanjidata[kanji.index(kanji2)][1]
                onyomi2 = kanjidata[kanji.index(kanji2)][2]
            except:
                kanji2 = ""
                meaning2 = ""
                kunyomi2 = ""
                onyomi2 = ""
            try:
                kanji3 = kanjiused[2]
                meaning3 = kanjidata[kanji.index(kanji3)][0]
                kunyomi3 = kanjidata[kanji.index(kanji3)][1]
                onyomi3 = kanjidata[kanji.index(kanji3)][2]
            except:
                kanji3 = ""
                meaning3 = ""
                kunyomi3 = ""
                onyomi3 = ""
            try:
                kanji4 = kanjiused[3]
                meaning4 = kanjidata[kanji.index(kanji4)][0]
                kunyomi4 = kanjidata[kanji.index(kanji4)][1]
                onyomi4 = kanjidata[kanji.index(kanji4)][2]
            except:
                kanji4 = ""
                meaning4 = ""
                kunyomi4 = ""
                onyomi4 = ""

            furigana = getFurigana(word, reading)
            partsofspeech = getPartOfSpeech(j)
            tags = [getJLPT(j),getCommonality(j), partsofspeech]

            ws.append([furigana, meaning, " ".join(tags), partsofspeech.replace(' ',', '), kanji1, meaning1, onyomi1, kunyomi1, kanji2, meaning2, onyomi2, kunyomi2, kanji3, meaning3, onyomi3, kunyomi3, kanji4, meaning4, onyomi4, kunyomi4])

def kanjiSearch(kanji):
        url = requests.get("https://jisho.org/search/" + str(kanji) + "%20%23kanji")
        soup = BeautifulSoup(url.content, 'html.parser')

        meaning = ""
        kunyomi = ""
        onyomi = ""
        frequency = ""
        #Kanji Meaning
        if soup.find_all("div", {"class": "kanji-details__main-meanings"}):
            meaning = soup.find_all("div", {"class": "kanji-details__main-meanings"})[0].text.strip()

        #Kunyomi
        try:
            if soup.find_all("dd", {"class": "kanji-details__main-readings-list"}):
                kunyomi = soup.find_all("dd", {"class": "kanji-details__main-readings-list"})[0].text.strip()
        except:
            pass
        #Onyomi
        try:
            if soup.find_all("dd", {"class": "kanji-details__main-readings-list"}):
                onyomi = soup.find_all("dd", {"class": "kanji-details__main-readings-list"})[1].text.strip()
        except:
            pass
        
        #Newspaper Frequency
        if soup.find_all("div", {"class": "frequency"}):
            frequency = soup.find_all("div", {"class": "frequency"})[0].text.strip()

        info = [meaning, kunyomi, onyomi ,frequency]
        return info

def kanjiSearchToExcel():
    
    for i in kanjilist:
        url = requests.get("https://jisho.org/search/" + str(i) + "%20%23kanji")
        soup = BeautifulSoup(url.content, 'html.parser')

        #Newspaper Frequency
        if soup.find_all("div", {"class": "frequency"}):
            frequency = soup.find_all("div", {"class": "frequency"})[0].text.strip()
        #Kanji Meaning
        if soup.find_all("div", {"class": "kanji-details__main-meanings"}):
            meaning = soup.find_all("div", {"class": "kanji-details__main-meanings"})[0].text.strip()
        #Kunyomi
        try:
            if soup.find_all("dd", {"class": "kanji-details__main-readings-list"}):
                kunyomi = soup.find_all("dd", {"class": "kanji-details__main-readings-list"})[0].text.strip()
        except:
            pass
        #Onyomi
        try:
            if soup.find_all("dd", {"class": "kanji-details__main-readings-list"}):
                onyomi = soup.find_all("dd", {"class": "kanji-details__main-readings-list"})[1].text.strip()
        except:
            pass
        
        #Grade
        try:
            if soup.find_all("div", {"class": "grade"}):
                grade = soup.find_all("div", {"class": "grade"})[0].text.strip()
                grade = grade[grade.find('grade'):].replace(' ','-')
        except:
            pass

        #JLPT
        try:
            if soup.find_all("div", {"class": "jlpt"}):
                jlpt = soup.find_all("div", {"class": "jlpt"})[0].text.strip()
                jlpt = "jlpt-" + jlpt[11:].lower()
        except:
            pass

        tags = jlpt + " " + grade
        words = readPage(1, 1, "*" + urllib.parse.quote(str(i) + "*"))
        definitions = ""
        index = 0
        for j in words:
                for k in j['data']:
                    if index < 5:
                        defword = getWord(k)
                        defreading = getReading(k)
                        defmeaning = getMeaning(k, 1, True)
                        deffurigana = getFurigana(defword, defreading)
                        definitions = definitions + '"' + deffurigana + " " + defmeaning + '"' + " &CHAR(10)& "
                        index += 1
        definitions = "= " + definitions[:-12]
        ws2.append([i, meaning, frequency, kunyomi, onyomi, definitions, tags])

def kanjiSearchFromExcelWords():
    for i in kanjilist:
        

wordSearchToExcel(1, 1000, n4)
kanjiSearchToExcel()

#Save the file
wb.save("sample.xlsx")